from django.shortcuts import get_object_or_404, redirect, render
from django.http import HttpResponse
import yaml

from orders.settings import EMAIL_HOST_USER
from .forms import UserRegistrationForm
from .forms import ShopRegistrationForm
from .models import Order, OrderItem, Product, Category, ProductInfo, Shop
from django.core.mail import send_mail
from django.db.models import Q, Min
from django.http import HttpResponseRedirect
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from django.contrib import messages
from django.core.mail import EmailMessage
from django.template.loader import get_template
from django.template import Context
from django.conf import settings
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from django.contrib.auth.models import User
import openpyxl
from openpyxl import Workbook

# Create your views here.
  
def index(request):
    return render(request, 'backend/index.html')

def products_list(request):
    products = Product.objects.all()

    # Получаем значения фильтров из GET-параметров
    name = request.GET.get('name')
    category = request.GET.get('category')
    min_price = request.GET.get('min_price')
    max_price = request.GET.get('max_price')

    # Применяем фильтры
    if name:
        products = products.filter(name__icontains=name)
    if category:
        products = products.filter(category__name=category)
    if min_price:
        products = products.filter(product_infos__price__gte=min_price)
    if max_price:
        products = products.filter(product_infos__price__lte=max_price)

    products = products.annotate(min_price=Min('product_infos__price'))

    context = {
        'products': products
    }
    return render(request, 'backend/products_list.html', context)

def all_categories(request):
    categories_list = Category.objects.all()
    return render(request, 'backend/categories_list.html', {'categories_list': categories_list})

def all_shops(request):
    shops_list = Shop.objects.all()
    return render(request, 'backend/shops_list.html', {'shops_list': shops_list})

def product_detail(request, product_id):
    product = get_object_or_404(Product, pk=product_id)
    product_infos = product.product_infos.all()

    context = {
        'product': product,
        'product_infos': product_infos
    }
    return render(request, 'backend/product_detail.html', context)

def add_to_cart(request, product_info_id):
    if request.method == 'POST':
        product_info = get_object_or_404(ProductInfo, id=product_info_id)
        quantity = request.POST.get('quantity')
        
        try:
            order = Order.objects.get(user=request.user, status='editing')
            order.add_product(product_info, quantity)
        except Order.DoesNotExist:
            order = Order.objects.create(user=request.user, status='editing')
            order.add_product(product_info, quantity)

        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    
def orders(request):
    user = request.user
    orders = Order.objects.filter(user=user)
    context = {
        'orders': orders
    }
    return render(request, 'backend/orders.html', context)

def cart(request):
    user = request.user
    order = Order.objects.filter(user=user, status='editing').first()
    
    if order:
        order_items = order.orderitem_set.all()
        total_price = sum(item.quantity * item.price for item in order_items)
    else:
        order_items = {}
        total_price = 0
    
    if request.method == 'POST':
        delivery_address = request.POST.get('delivery_address')
        delivery_date = request.POST.get('delivery_date')

        order.delivery_address = delivery_address
        order.delivery_date = delivery_date
        order.status = 'confirmed'
        order.save()

        # отправка email

        # создание PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="order_confirmation.pdf"'
        
        pdfmetrics.registerFont(TTFont('ArialUnicode', 'backend/fonts/ArialUnicode.ttf'))

        doc = SimpleDocTemplate(response, pagesize=letter)
        elements = []

        styles = {
            'Heading1': ParagraphStyle(
                'Heading1',
                fontSize=16,
                fontName='ArialUnicode',
                spaceAfter=12
            ),
            'Heading2': ParagraphStyle(
                'Heading2',
                fontSize=14,
                fontName='ArialUnicode',
                spaceAfter=6
            ),
        }

        # Добавляем детали заказа
        elements.append(Paragraph('Ваш заказ:', styles['Heading1']))
        elements.append(Paragraph('', styles['Heading1']))
        elements.append(Paragraph(f'ID заказа: {order.id}', styles['Heading2']))
        elements.append(Paragraph(f'ID получателя: {user.id}', styles['Heading2']))
        elements.append(Paragraph('', styles['Heading1']))        
        elements.append(Paragraph(f'Получатель: {user.first_name} {user.last_name}', styles['Heading2']))
        elements.append(Paragraph(f'Адрес доставки: {order.delivery_address}', styles['Heading2']))
        elements.append(Paragraph(f'Ожидаемая дата доставки: {order.delivery_date}', styles['Heading2']))
        elements.append(Paragraph(f'Общая сумма заказа: {total_price}', styles['Heading2']))
        elements.append(Paragraph('', styles['Heading1']))
        # Добавляем позиции по заказу
        data = [['Наименование', 'Количество', 'Цена', 'Дистрибъютор']]
        for item in order_items:
            data.append([item.product_name, 
                         item.quantity, 
                         item.price, 
                         item.shop])

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'ArialUnicode'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('FONTNAME', (0, 1), (-1, -1), 'ArialUnicode'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))

        elements.append(table)
        doc.build(elements)

        # отправка email
        email = EmailMessage(
            subject='Ваш заказ оформлен',
            body=f'Ваш заказ по адресу будет доставлен {delivery_date} по адресу {delivery_address}. Сумма заказа: {total_price}',
            from_email=EMAIL_HOST_USER,
            to=[user.email],
        )
        email.attach('order_confirmation.pdf', response.getvalue(), 'application/pdf')
        email.send()

        order_items = order.orderitem_set.all()
        for item in order_items:
            shop = item.shop
            manager_emails = User.objects.filter(shop__name=shop).values_list('email', flat=True)
            
            pdf_data = generate_pdf(order, order_items, shop)
            
            email_subject = 'Заказ с сайте'
            email_body = 'Здравствуйте,\n\nВаш заказ с сайта:'
            email = EmailMessage(email_subject, email_body, settings.DEFAULT_FROM_EMAIL, manager_emails)
            email.attach('order.pdf', pdf_data, 'application/pdf')
            email.send()

        return redirect('thank_you_page')

    context = {
        'order': order,
        'order_items': order_items,
        'total_price': total_price
    }
    return render(request, 'backend/cart.html', context)

def delete_order_item(request, item_id):
    item = OrderItem.objects.get(id=item_id)
    item.delete()
    return redirect('cart')

def update_order_item(request, item_id):
    item = get_object_or_404(OrderItem, id=item_id)
    if request.method == 'POST':
        quantity = request.POST.get('quantity')
        item.quantity = quantity
        item.save()
    return redirect('cart')

def order_detail(request, order_id):
    order = Order.objects.get(id=order_id)
    order_items = order.orderitem_set.all()
    total_price = sum(item.quantity * item.price for item in order_items)
    context = {
        'order': order,
        'order_items': order_items,
        'total_price': total_price
    }
    return render(request, 'backend/order_detail.html', context)

def thank_you_page(request):
    user = request.user

    context = {
        'user': user
    }

    return render(request, 'backend/thank_you_page.html', context)

def register_shop(request):
    if request.method == 'POST':
        form = ShopRegistrationForm(request.POST)
        if form.is_valid():
            user = request.user
            if not Shop.objects.filter(user=user).exists():
                shop = form.save(commit=False)
                shop.user = request.user
                shop.save()
                messages.success(request, 'Магазин зарегистрирован.')
                return redirect('home')
            else:
                form.add_error(None, "Вы уже являетесь менеджером другого магазина")
    else:
        form = ShopRegistrationForm()

    return render(request, 'backend/register_shop.html', {'form': form})

def shop_catalog(request):
    user = request.user
    shop = Shop.objects.filter(user=user).first()
    
    products = ProductInfo.objects.filter(shop=shop).all()

    context = {
        'products':products,
        'shop':shop
    }
    return render(request, 'backend/shop_catalog.html', context)

def shop_status(request, shop_id):
    shop = get_object_or_404(Shop, id=shop_id)

    if request.method == 'POST':
        state = request.POST.get('state')
        shop.state = bool(state)
        shop.save()
        messages.success(request, 'Статус магазина изменен')
        return redirect('shop_catalog')

    return render(request, 'shop_catalog.html', {'shop': shop})

def import_products(request):
    if request.method == 'POST':
        yaml_file = request.FILES['yaml_file']
        try:
            data = yaml.safe_load(yaml_file)
        except yaml.YAMLError:
            messages.error(request, 'Ошибка чтения файла YAML')
            return redirect('import_products')

        shop_id = data.get('shop_id')
        shop = Shop.objects.get(pk=shop_id)
        user = request.user

        if shop.user == user: # Проверка соответствия id магазина с id пользователя
            products = data.get('products', [])
            ProductInfo.objects.filter(shop=shop).delete()
            for product_data in products:
                product = Product.objects.create(
                    name=product_data.get('name'),
                    category_id=product_data.get('category_id')
                )
                ProductInfo.objects.create(
                    external_id=product_data.get('external_id'),
                    product=product,
                    shop=shop,
                    quantity=product_data.get('quantity'),
                    price=product_data.get('price'),
                    price_rrc=product_data.get('price_rrc')
                )
            messages.success(request, 'Продукты успешно импортированы')
        else:
            messages.error(request, 'Ошибка: Вы не можете импортировать продукты для этого магазина')

        return redirect('shop_catalog')

    return render(request, 'backend/import.html')

def generate_pdf(order, order_items, shop):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)

    p.setFont("ArialUnicode", 14)
    p.drawString(30, 750, f"ID заказа: {order.id}")
    p.drawString(30, 725, f"Дата заказа: {order.delivery_date}")
    p.drawString(30, 700, "Адрес доставки: {address}".format(address=order.delivery_address))
    p.drawString(30, 675, f"Дата доставки: {order.delivery_date}")

    p.setFont("ArialUnicode", 12)
    p.drawString(30, 650, "Список товаров:")
    p.setFont("ArialUnicode", 10)
    
    y = 625
    order_summ = 0
    for item in order_items:
        if item.shop == shop:
            text = "{name} - {quantity} шт. - {price} руб.".format(name=item.product_name, quantity=item.quantity, price=item.price)
            p.drawString(30, y, text)
            order_summ += item.price * item.quantity
            y -= 25

    y -= 25
    p.drawString(30, y, f"Сумма заказа: {order_summ}")

    p.showPage()
    p.save()

    buffer.seek(0)
    return buffer.getvalue()

def export_orders(request, shop_id):
    shop = get_object_or_404(Shop, id=shop_id)
    orders = Order.objects.filter(orderitem__shop=shop).distinct()

    # Создаем новую рабочую книгу
    wb = openpyxl.Workbook()
    # Выбираем активный лист
    ws = wb.active

    # Заголовки столбцов
    ws['A1'] = 'Номер заказа'
    ws['B1'] = 'Дата создания'
    ws['C1'] = 'Адрес доставки'
    ws['D1'] = 'Дата доставки'
    ws['E1'] = 'Статус'
    ws['F1'] = 'Название товара'
    ws['G1'] = 'Кол-во'
    ws['H1'] = 'Цена'

    row_num = 2  # Номер строки для данных

    # Заполняем данные по заказам
    for order in orders:
        order_items = OrderItem.objects.filter(order=order, shop=shop)
        for item in order_items:
            ws.cell(row=row_num, column=1, value=order.id)
            ws.cell(row=row_num, column=2, value=order.order_date.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=3, value=order.delivery_address)
            ws.cell(row=row_num, column=4, value=order.delivery_date.strftime('%Y-%m-%d') if order.delivery_date else '')
            ws.cell(row=row_num, column=5, value=order.get_status_display())
            ws.cell(row=row_num, column=6, value=item.product_name)
            ws.cell(row=row_num, column=7, value=item.quantity)
            ws.cell(row=row_num, column=8, value=item.price)
            row_num += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=orders_{shop.name}.xlsx'
    wb.save(response)

    return response

# def export_orders(request, shop_id):
#     # Получаем объект магазина по его ID
#     shop = Shop.objects.get(id=shop_id)
    
#     # Получаем все заказы, связанные с данным магазином
#     orders = Order.objects.filter(orderitem__shop=shop).distinct()
    
#     # Создаем новый Excel-документ
#     workbook = Workbook()
#     sheet = workbook.active
    
#     # Заполняем заголовки столбцов
#     sheet['A1'] = 'Номер заказа'
#     sheet['B1'] = 'Адрес доставки'
#     sheet['C1'] = 'Дата доставки'
#     sheet['D1'] = 'Дата заказа'
#     sheet['E1'] = 'Название товара'
#     sheet['F1'] = 'Количество'
#     sheet['G1'] = 'Цена'
    
#     # Заполняем данные по заказам и продуктам
#     for row_num, order in enumerate(orders, start=2):
#         sheet.cell(row=row_num, column=1, value=order.id)
#         sheet.cell(row=row_num, column=2, value=order.delivery_address)
#         sheet.cell(row=row_num, column=3, value=order.delivery_date)
#         sheet.cell(row=row_num, column=4, value=order.order_date)
        
#         order_items = OrderItem.objects.filter(order=order, shop=shop)
#         for item in order_items:
#             sheet.cell(row=row_num, column=5, value=item.product_name)
#             sheet.cell(row=row_num, column=6, value=item.quantity)
#             sheet.cell(row=row_num, column=7, value=item.price)
#             row_num += 1
    
#     # Настраиваем HTTP-response для скачивания файла
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename=orders.xlsx'
    
#     # Сохраняем Excel-документ в HTTP-response
#     workbook.save(response)
    
#     return response